import os;
import numpy as np;
import matplotlib.pyplot as plt;
from resultUtils import matrixAccuracy;

def autolabel(rects, plt, means, maxValue):
    """
    Attach a text label above each bar displaying its height
    """
    if maxValue > 70:
        offset = 5.7;
        increase1 = 5;
        increase2 = -3;
    elif maxValue > 50:
        offset = 4.5;
        increase1 = 5;
        increase2 = 0;
    else:
        offset = 3;
        increase1 = -5;
        increase2 = -3;
    i = 0;
    for rect in rects:
        height = rect.get_height();
        if 0:
            if i == 0 and height < 11:
                textHeight = 12+increase1;
            if i == 0 and height < 15:
                textHeight = 18+increase1;
            elif i == 0:
                textHeight = 10+increase2;
            else:
                textHeight = 10+increase2;
            if height > 12:
                plt.text(rect.get_x() + rect.get_width()/2., textHeight-6,'%3.1f%c' % (height, '%'), ha='center', va='bottom', fontsize=14)
                #plt.text(rect.get_x() + rect.get_width()/2., textHeight-offset,'+-%3.1f' % (means[1][i]), ha='center', va='bottom', fontsize=14)
            else:
                plt.text(rect.get_x() + rect.get_width()/2., textHeight-8,'%3.1f%c' % (height, '%'), ha='center', va='bottom', fontsize=14)
                #plt.text(rect.get_x() + rect.get_width()/2., textHeight-8-offset,'+-%3.1f' % (means[1][i]), ha='center', va='bottom', fontsize=14)
        else:
            plt.text(rect.get_x() + rect.get_width()/2., -9,'%3.1f%c' % (height, '%'), ha='center', va='bottom', fontsize=14)
        i += 1;     
        
def plotBarGraphCC(means, errors, title, filename):
    labels = ['Top 1', 'Top 5', 'Top 10', 'Kamitani Score'];

    barWidth = 0.15  # the width of the bars

    # Set position of bar on X axis
    r1 = np.arange(len(means[0]));
    r2 = [x + barWidth for x in r1];
    r3 = [x + barWidth for x in r2];
    r4 = [x + barWidth for x in r3];
    r5 = [x + barWidth for x in r4];
    r6 = [x + barWidth for x in r5];
    r7 = [x + barWidth for x in r6];

    plt.figure(figsize = (14, 4));
    # Make the plot
    plt.bar(r1, means[0], yerr=errors[0], width=barWidth, edgecolor='white', label='Single volume/200 classes/Mean prototype', align='center', alpha=0.5, ecolor='black', capsize=10);
    plt.bar(r2, means[1], yerr=errors[1], width=barWidth, edgecolor='white', label='Single volume/50 classes/Mean prototype',align='center', alpha=0.5, ecolor='black', capsize=10);
    plt.bar(r3, means[2], yerr=errors[2], width=barWidth, edgecolor='white', label='Mean volume/200 classes/Mean prototype',align='center', alpha=0.5, ecolor='black', capsize=10);
    plt.bar(r4, means[3], yerr=errors[3], width=barWidth, edgecolor='white', label='Mean volume/50 classes/Mean prototype',align='center', alpha=0.5, ecolor='black', capsize=10);
    plt.bar(r5, means[4], yerr=errors[4], width=barWidth, edgecolor='white', label='Single Volume/50 classes/Image prototype',align='center', alpha=0.5, ecolor='black', capsize=10);
    plt.bar(r6, means[5], yerr=errors[5], width=barWidth, edgecolor='white', label='Mean volume/50 classes/Image Prototype',align='center', alpha=0.5, ecolor='black', capsize=10);
    plt.bar(r7, means[6], yerr=errors[6], width=barWidth, edgecolor='white', label='[Horikawa 2017]',align='center', alpha=0.5, ecolor='black', capsize=10);
    # Add xticks on the middle of the group bars
    plt.xlabel('Accuracy scores', fontweight='bold')
    plt.xticks([r + barWidth for r in range(len(r1))], labels);
    plt.grid(True, axis='y');
     
    plt.ylabel('Accuracy');
    plt.title(title);
    
    # Create legend & Show graphic
    plt.legend();
    #plt.show();
    
    plt.savefig(filename);
   
def plotBarGraphCC2(means, errors, title, filename):
    barNames = ['Single volume', 'Mean   volume'];
    plotBarGraphDefault(means, errors, title, filename, barNames);

def plotBarGraphCC2A(means, errors, title, filename):
    labels = ['Top 1', 'Top 5', 'Top 10'];

    barWidth = 0.22;  # the width of the bars

    # Set position of bar on X axis
    r1 = np.arange(len(means[0]));
    r2 = [x + barWidth for x in r1];
    r3 = [x + barWidth + 0.0125 for x in r2];
    r4 = [x + barWidth for x in r3];

    plt.figure(figsize = (14, 4));
    # Make the plot
    b1 = plt.bar(r1, means[0], yerr=errors[0], width=barWidth, edgecolor='white', label='Single volume - fc7', align='center', alpha=0.5, ecolor='black');
    b2 = plt.bar(r2, means[1], yerr=errors[1], width=barWidth, edgecolor='white', label='Mean volume   - fc7',align='center', alpha=0.5, ecolor='black');
    b3 = plt.bar(r3, means[2], yerr=errors[2], width=barWidth, edgecolor='white', label='Single volume - conv5',align='center', alpha=0.5, ecolor='black');
    b4 = plt.bar(r4, means[3], yerr=errors[3], width=barWidth, edgecolor='white', label='Mean volume   - conv5',align='center', alpha=0.5, ecolor='black');

    # Add xticks on the middle of the group bars
    plt.xlabel('Accuracy scores', fontweight='bold', fontsize = 12);
    plt.xticks([r + barWidth for r in range(len(r1))], labels, fontsize = 12);
    plt.grid(True, axis='y');
     
    plt.ylabel('Accuracy', fontsize = 12);
    plt.title(title, fontsize = 12);
    
    # Create legend & Show graphic
    plt.legend(fontsize = 12);
    maxValue = 0;
    for i in range(4):
        for j in range(len(means[0])):
            maxValue = max(maxValue, means[i][j] + errors[i][1][j]);
    
    plt.ylim(bottom=-10, top=maxValue+10);
    # dotted lines
    plt.hlines(100/50, r1[0] - 0.7 * barWidth, r1[0] + 3.7*barWidth, colors='k', linestyles = 'dashed');
    plt.hlines(500/50, r1[1] - 0.7 * barWidth, r1[1] + 3.7*barWidth, colors='k', linestyles = 'dashed');
    plt.hlines(1000/50, r1[2] - 0.7 * barWidth, r1[2] + 3.7*barWidth, colors='k', linestyles = 'dashed');

    autolabel(b1, plt, errors[0], maxValue);
    autolabel(b2, plt, errors[1], maxValue);
    autolabel(b3, plt, errors[2], maxValue);
    autolabel(b4, plt, errors[3], maxValue);
    #plt.show();
    
    plt.savefig(filename);

def plotBarGraphCC3(means, errors, title, filename):
    barNames = ['Avg.  prototype', 'Image prototype'];
    plotBarGraphDefault(means, errors, title, filename, barNames);
    
def plotBarGraphCC3A(means, errors, title, filename):
    labels = ['Top 1', 'Top 5', 'Top 10'];

    barWidth = 0.22;  # the width of the bars

    # Set position of bar on X axis
    r1 = np.arange(len(means[0]));
    r2 = [x + barWidth for x in r1];
    r3 = [x + barWidth + 0.0125 for x in r2];
    r4 = [x + barWidth for x in r3];

    plt.figure(figsize = (14, 4));
    # Make the plot
    b1 = plt.bar(r1, means[0], yerr=errors[0], width=barWidth, edgecolor='white', label='Avg.  prototype - fc7', align='center', alpha=0.5, ecolor='black');
    b2 = plt.bar(r2, means[1], yerr=errors[1], width=barWidth, edgecolor='white', label='Image prototype - fc7',align='center', alpha=0.5, ecolor='black');
    b3 = plt.bar(r3, means[2], yerr=errors[2], width=barWidth, edgecolor='white', label='Avg.  prototype - conv5',align='center', alpha=0.5, ecolor='black');
    b4 = plt.bar(r4, means[3], yerr=errors[3], width=barWidth, edgecolor='white', label='Image prototype - conv5',align='center', alpha=0.5, ecolor='black');

    # Add xticks on the middle of the group bars
    plt.xlabel('Accuracy scores', fontweight='bold', fontsize = 12);
    plt.xticks([r + barWidth for r in range(len(r1))], labels, fontsize = 12);
    plt.grid(True, axis='y');
     
    plt.ylabel('Accuracy', fontsize = 12);
    plt.title(title, fontsize = 12);
    
    # Create legend & Show graphic
    plt.legend(fontsize = 12);
    maxValue = 0;
    for i in range(4):
        for j in range(len(means[0])):
            maxValue = max(maxValue, means[i][j] + errors[i][1][j]);
    
    plt.ylim(bottom=-10, top=maxValue+10);
    # dotted lines
    plt.hlines(100/50, r1[0] - 0.7 * barWidth, r1[0] + 3.7*barWidth, colors='k', linestyles = 'dashed');
    plt.hlines(500/50, r1[1] - 0.7 * barWidth, r1[1] + 3.7*barWidth, colors='k', linestyles = 'dashed');
    plt.hlines(1000/50, r1[2] - 0.7 * barWidth, r1[2] + 3.7*barWidth, colors='k', linestyles = 'dashed');

    autolabel(b1, plt, errors[0], maxValue);
    autolabel(b2, plt, errors[1], maxValue);
    autolabel(b3, plt, errors[2], maxValue);
    autolabel(b4, plt, errors[3], maxValue);
    
    #plt.show();
    
    plt.savefig(filename);

def plotBarGraphNN(means, errors, title, filename):
    labels = ['Top 1', 'Top 5', 'Top 10', 'Kamitani Score'];

    barWidth = 0.20  # the width of the bars

    # Set position of bar on X axis
    r1 = np.arange(len(means[0]));
    r2 = [x + barWidth for x in r1];
    r3 = [x + barWidth for x in r2];
    r4 = [x + barWidth for x in r3];

    #plt.figure(figsize = (10, 8));
    # Make the plot
    b1 = plt.bar(r1, means[0], yerr=errors[0], width=barWidth, edgecolor='white', label='Single volume/200 classes', align='center', alpha=0.5, ecolor='black');
    b2 = plt.bar(r2, means[1], yerr=errors[1], width=barWidth, edgecolor='white', label='Single volume/50 classes',align='center', alpha=0.5, ecolor='black');
    b3 = plt.bar(r3, means[2], yerr=errors[2], width=barWidth, edgecolor='white', label='Mean volume/200 classes',align='center', alpha=0.5, ecolor='black');
    b4 = plt.bar(r4, means[3], yerr=errors[3], width=barWidth, edgecolor='white', label='Mean volume/50 classes',align='center', alpha=0.5, ecolor='black');
      
    # Add xticks on the middle of the group bars
    plt.xlabel('Accuracy scores', fontweight='bold')
    plt.xticks([r + barWidth for r in range(len(r1))], labels);
    plt.grid(True, axis='y');
     
    plt.ylabel('Accuracy');
    plt.title(title);
    
    # Create legend & Show graphic
    plt.legend();
    
    #plt.show();
    
    plt.savefig(filename);

def plotBarGraphNN2A(means, errors, title, filename):
    labels = ['Top 1', 'Top 5', 'Top 10'];

    barWidth = 0.22  # the width of the bars

    plt.figure(figsize = (14, 4));
    # Set position of bar on X axis
    r1 = np.arange(len(means[0]));
    r2 = [x + barWidth for x in r1];
    r3 = [x + barWidth + 0.0125 for x in r2];
    r4 = [x + barWidth for x in r3];

    #plt.figure(figsize = (10, 8));
    # Make the plot
    b1 = plt.bar(r1, means[0], yerr=errors[0], width=barWidth, edgecolor='white', label='Single volume - fc7', align='center', alpha=0.5, ecolor='black', capsize=0);
    b2 = plt.bar(r2, means[1], yerr=errors[1], width=barWidth, edgecolor='white', label='Mean   volume - fc7',align='center', alpha=0.5, ecolor='black', capsize=0);
    b3 = plt.bar(r3, means[2], yerr=errors[2], width=barWidth, edgecolor='white', label='Single volume - conv5',align='center', alpha=0.5, ecolor='black', capsize=0);
    b4 = plt.bar(r4, means[3], yerr=errors[3], width=barWidth, edgecolor='white', label='Mean   volume - conv5',align='center', alpha=0.5, ecolor='black', capsize=0);
      
    # Add xticks on the middle of the group bars
    plt.xlabel('Accuracy scores', fontsize = 12, fontweight='bold')
    plt.xticks([r + 1.5*barWidth for r in range(len(r1))], labels, fontsize=12);
    plt.grid(True, axis='y');
     
    plt.ylabel('Accuracy', fontsize=12);
    plt.title(title, fontsize = 14);
    
    # Create legend & Show graphic
    plt.legend(loc = 'upper left', fontsize=12);

    maxValue = 0;
    for i in range(4):
        for j in range(len(means[0])):
            maxValue = max(maxValue, means[i][j] + errors[i][1][j]);
    
    plt.ylim(bottom=-10, top=maxValue+10);
    # dotted lines
    plt.hlines(100/50, r1[0] - 0.7 * barWidth, r1[0] + 3.7*barWidth, colors='k', linestyles = 'dashed');
    plt.hlines(500/50, r1[1] - 0.7 * barWidth, r1[1] + 3.7*barWidth, colors='k', linestyles = 'dashed');
    plt.hlines(1000/50, r1[2] - 0.7 * barWidth, r1[2] + 3.7*barWidth, colors='k', linestyles = 'dashed');
    
    autolabel(b1, plt, errors[0], maxValue);
    autolabel(b2, plt, errors[1], maxValue);
    autolabel(b3, plt, errors[2], maxValue);
    autolabel(b4, plt, errors[3], maxValue);

    #plt.show();
    
    plt.savefig(filename);


def plotBarGraphDefault(means, errors, title, filename, barNames):
    imagined = 'Imagined' in title;
    
    labels = ['Top 1', 'Top 5', 'Top 10'];

    barWidth = 0.4  # the width of the bars

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(15, 5));
    fig.subplots_adjust(top=0.85, bottom=0.1, wspace=0.0)
    fig.suptitle(title, fontsize = 14);
    # Set position of bar on X axis
    r1 = np.arange(len(means[0]));
    r2 = [x + barWidth for x in r1];

    maxValue = 0;
    for i in range(4):
        for j in range(len(means[0])):
            maxValue = max(maxValue, means[i][j] + errors[i][1][j]);

    #plt.figure(figsize = (10, 8));
    # Make the plot
    
    prop_cycle = plt.rcParams['axes.prop_cycle'];
    colors = prop_cycle.by_key()['color'];
    ax2.get_yaxis().set_ticklabels([]);

    if imagined:
        c1 = colors[4];
        c2 = colors[5];
    else:
        c1 = colors[0];
        c2 = colors[1];
        
    b1 = ax1.bar(r1, means[0], yerr=errors[0], width=barWidth, edgecolor='white', label=barNames[0], align='center', alpha=0.5, ecolor='black', capsize=0, color=c1);
    b2 = ax1.bar(r2, means[1], yerr=errors[1], width=barWidth, edgecolor='white', label=barNames[1], align='center', alpha=0.5, ecolor='black', capsize=0, color=c2);

    autolabel(b1, ax1, errors[0], maxValue);
    autolabel(b2, ax1, errors[1], maxValue);
    
    b3 = ax2.bar(r1, means[2], yerr=errors[2], width=barWidth, edgecolor='white', label=barNames[0], align='center', alpha=0.5, ecolor='black', capsize=0, color=c1);
    b4 = ax2.bar(r2, means[3], yerr=errors[3], width=barWidth, edgecolor='white', label=barNames[1], align='center', alpha=0.5, ecolor='black', capsize=0, color=c2);

    autolabel(b3, ax2, errors[2], maxValue);
    autolabel(b4, ax2, errors[3], maxValue);

      
    # Add xticks on the middle of the group bars
    ax1.set_title('FC7', fontsize = 15, fontweight='bold');
    ax2.set_title('Conv5', fontsize = 15, fontweight='bold');
    
    ax1.set_xticks([r + .5*barWidth for r in range(len(r1))]);
    ax1.set_xticklabels(labels, fontsize=15);
    ax1.grid(True, axis='y');

    ax2.set_xticks([r + .5*barWidth for r in range(len(r1))]);
    ax2.set_xticklabels( labels, fontsize=15);
    ax2.grid(True, axis='y');
     
    ax1.set_ylabel('Accuracy', fontsize=15);
    
    # Create legend & Show graphic
    ax1.legend(loc = 'upper left', fontsize=15);

    ax1.set_ylim(bottom=-10, top=maxValue+10);
    ax2.set_ylim(bottom=-10, top=maxValue+10);
    
    # dotted lines
    ax1.hlines(100/50, r1[0] - 0.7 * barWidth, r1[0] + 1.7*barWidth, colors='k', linestyles = 'dashed');
    ax1.hlines(500/50, r1[1] - 0.7 * barWidth, r1[1] + 1.7*barWidth, colors='k', linestyles = 'dashed');
    ax1.hlines(1000/50, r1[2] - 0.7 * barWidth, r1[2] + 1.7*barWidth, colors='k', linestyles = 'dashed');

    ax2.hlines(100/50, r1[0] - 0.7 * barWidth, r1[0] + 1.7*barWidth, colors='k', linestyles = 'dashed');
    ax2.hlines(500/50, r1[1] - 0.7 * barWidth, r1[1] + 1.7*barWidth, colors='k', linestyles = 'dashed');
    ax2.hlines(1000/50, r1[2] - 0.7 * barWidth, r1[2] + 1.7*barWidth, colors='k', linestyles = 'dashed');
    
    plt.savefig(filename);

def plotBarGraphNN2(means, errors, title, filename):
    barNames = ['Single volume', 'Mean   volume'];
    plotBarGraphDefault(means, errors, title, filename, barNames);

def plotBarGraphNN3A(means, errors, title, filename):
    labels = ['Top 1', 'Top 5', 'Top 10'];

    barWidth = 0.22  # the width of the bars

    plt.figure(figsize = (14, 4));
    # Set position of bar on X axis
    r1 = np.arange(len(means[0]));
    r2 = [x + barWidth for x in r1];
    r3 = [x + barWidth + 0.0125 for x in r2];
    r4 = [x + barWidth for x in r3];

    #plt.figure(figsize = (10, 8));
    # Make the plot
    b1 = plt.bar(r1, means[0], yerr=errors[0], width=barWidth, edgecolor='white', label='Neural netwroks  - fc7', align='center', alpha=0.5, ecolor='black', capsize=0);
    b2 = plt.bar(r2, means[1], yerr=errors[1], width=barWidth, edgecolor='white', label='Correlation rank - fc7',align='center', alpha=0.5, ecolor='black', capsize=0);
    b3 = plt.bar(r3, means[2], yerr=errors[2], width=barWidth, edgecolor='white', label='Neural networks  - conv5',align='center', alpha=0.5, ecolor='black', capsize=0);
    b4 = plt.bar(r4, means[3], yerr=errors[3], width=barWidth, edgecolor='white', label='Correlation Rank - conv5',align='center', alpha=0.5, ecolor='black', capsize=0);
      
    # Add xticks on the middle of the group bars
    plt.xlabel('Accuracy scores', fontsize = 12, fontweight='bold')
    plt.xticks([r + 1.5*barWidth for r in range(len(r1))], labels, fontsize=12);
    plt.grid(True, axis='y');
     
    plt.ylabel('Accuracy', fontsize=12);
    plt.title(title, fontsize = 14);
    
    # Create legend & Show graphic
    plt.legend(loc = 'upper left', fontsize=12);

    maxValue = 0;
    for i in range(4):
        for j in range(len(means[0])):
            maxValue = max(maxValue, means[i][j] + errors[i][1][j]);
    
    plt.ylim(bottom=-10, top=maxValue+10);
    # dotted lines
    plt.hlines(100/50, r1[0] - 0.7 * barWidth, r1[0] + 3.7*barWidth, colors='k', linestyles = 'dashed');
    plt.hlines(500/50, r1[1] - 0.7 * barWidth, r1[1] + 3.7*barWidth, colors='k', linestyles = 'dashed');
    plt.hlines(1000/50, r1[2] - 0.7 * barWidth, r1[2] + 3.7*barWidth, colors='k', linestyles = 'dashed');
    
    autolabel(b1, plt, errors[0], maxValue);
    autolabel(b2, plt, errors[1], maxValue);
    autolabel(b3, plt, errors[2], maxValue);
    autolabel(b4, plt, errors[3], maxValue);

    #plt.show();
    
    plt.savefig(filename);

def plotBarGraphNN3(means, errors, title, filename):
    barNames = ['Neural Networks', 'Correlation Rank'];
    plotBarGraphDefault(means, errors, title, filename, barNames);

def plotBarGraphHuge(means, errors, title, filename):
    labels = ['Top 1', 'Top 5', 'Top 10', 'Kamitani Score'];

    barWidth = 0.20  # the width of the bars

    # Set position of bar on X axis
    r1 = np.arange(len(means[0]));
    r2 = [x + barWidth for x in r1];
    r3 = [x + barWidth for x in r2];

    #plt.figure(figsize = (10, 8));
    # Make the plot
    plt.bar(r1, means[0], yerr=errors[0], width=barWidth, edgecolor='white', label='Single volume/15225 classes', align='center', alpha=0.5, ecolor='black', capsize=10);
    plt.bar(r2, means[1], yerr=errors[1], width=barWidth, edgecolor='white', label='Mean volume/15225 classes',align='center', alpha=0.5, ecolor='black', capsize=10);
    plt.bar(r3, means[2], yerr=errors[2], width=barWidth, edgecolor='white', label='[Horikawa 2017]',align='center', alpha=0.5, ecolor='black', capsize=10);
      
    # Add xticks on the middle of the group bars
    plt.xlabel('Accuracy scores', fontweight='bold')
    plt.xticks([r + barWidth for r in range(len(r1))], labels);
    plt.grid(True, axis='y');
     
    plt.ylabel('Accuracy');
    plt.title(title);
    
    # Create legend & Show graphic
    plt.legend();
    #plt.show();
    
    plt.savefig(filename);

def plotBarGraphHuge2(means, errors, title, filename):
    imagined = 'Imagined' in title;

    labels = ['FC7', 'Conv5'];

    barWidth = 0.22;  # the width of the bars

    plt.figure(figsize = (8, 4));

    # Set position of bar on X axis
    r1 = np.arange(len(means[0]));
    r2 = [x + barWidth for x in r1];

    #plt.figure(figsize = (10, 8));
    # Make the plot
    prop_cycle = plt.rcParams['axes.prop_cycle'];
    colors = prop_cycle.by_key()['color'];

    if imagined:
        c1 = colors[4];
        c2 = colors[5];
    else:
        c1 = colors[0];
        c2 = colors[1];

    b1 = plt.bar(r1, means[0], yerr=errors[0], width=barWidth, edgecolor='white', label='Our result',align='center', alpha=0.5, ecolor='black', color=c1);
    b2 = plt.bar(r2, means[1], yerr=errors[1], width=barWidth, edgecolor='white', label='[Horikawa 2017]',align='center', alpha=0.5, ecolor='black', color=c2);
      
    # Add xticks on the middle of the group bars
    plt.xlabel('Accuracy scores', fontweight='bold', fontsize = 12)
    plt.xticks([r + barWidth for r in range(len(r1))], labels, fontsize = 12);
    plt.grid(True, axis='y');
     
    plt.ylabel('Accuracy', fontsize = 12);
    plt.title(title, fontsize = 12);
    
    # Create legend & Show graphic
    plt.legend(loc = 'lower center', fontsize = 12);
    
    plt.ylim(bottom=-10, top=100);
    plt.yticks(np.arange(0, 100, 25.0));
    # dotted lines
    plt.hlines(50, r1[0] - 0.7 * barWidth, r1[0] + 1.7*barWidth, colors='k', linestyles = 'dashed');
    plt.hlines(50, r1[1] - 0.7 * barWidth, r1[1] + 1.7*barWidth, colors='k', linestyles = 'dashed');

    maxValue = 0;
    for i in range(2):
        for j in range(len(means[0])):
            maxValue = max(maxValue, means[i][j] + errors[i][1][j]);

    autolabel(b1, plt, errors[0], maxValue);
    autolabel(b2, plt, errors[1], maxValue);
    #plt.show();
    
    plt.savefig(filename);
    
def plotBarGraphHuge3(means, errors, title, filename):
    labels = ['FC7', 'Conv5'];

    barWidth = 0.20  # the width of the bars

    # Set position of bar on X axis
    r1 = np.arange(len(means[0]));
    r2 = [x + barWidth for x in r1];
    r3 = [x + barWidth for x in r2];

    #plt.figure(figsize = (10, 8));
    # Make the plot
    plt.bar(r1, means[0], yerr=errors[0], width=barWidth, edgecolor='white', label='Single volume', align='center', alpha=0.5, ecolor='black', capsize=10);
    plt.bar(r2, means[1], yerr=errors[1], width=barWidth, edgecolor='white', label='Mean volume',align='center', alpha=0.5, ecolor='black', capsize=10);
    plt.bar(r3, means[2], yerr=errors[2], width=barWidth, edgecolor='white', label='[Horikawa 2017]',align='center', alpha=0.5, ecolor='black', capsize=10);
      
    # Add xticks on the middle of the group bars
    plt.xlabel('Accuracy scores', fontweight='bold')
    plt.xticks([r + barWidth for r in range(len(r1))], labels);
    plt.grid(True, axis='y');
     
    plt.ylabel('Accuracy');
    plt.title(title);
    
    # Create legend & Show graphic
    plt.legend();
    #plt.show();
    
    plt.savefig(filename);


def plotBarGraphMulti(means, errors, title, filename):
    labels = ['Top 1', 'Top 5', 'Top 10', 'Kamitani Score'];

    barWidth = 0.20  # the width of the bars

    # Set position of bar on X axis
    r1 = np.arange(len(means[0]));
    r2 = [x + barWidth for x in r1];

    plt.figure(figsize = (10, 8));
    # Make the plot
    plt.bar(r1, means[0], yerr=errors[0], width=barWidth, edgecolor='white', label='200 classes/single data', align='center', alpha=0.5, ecolor='black', capsize=10);
    plt.bar(r2, means[1], yerr=errors[1], width=barWidth, edgecolor='white', label='150 classes Leave one out/single data',align='center', alpha=0.5, ecolor='black', capsize=10);
      
    # Add xticks on the middle of the group bars
    plt.xlabel('Accuracy scores', fontweight='bold')
    plt.xticks([r + barWidth for r in range(len(r1))], labels);
    plt.grid(True, axis='y');
     
    plt.ylabel('Accuracy');
    plt.title(title);
    
    # Create legend & Show graphic
    plt.legend();
    #plt.show();
    
    plt.savefig(filename);

def normalizeAccuracies(acc, numClasses):
    chance = 100.0 / numClasses;
    return 100.0 * (acc-chance) / (100.0-chance);
       
def plotResultsCC(excelfile, outputfile, layer, seenAcc, imagAcc):
    import xlrd;
    
    wb = xlrd.open_workbook(excelfile); 

    barsSeen = np.zeros((7,4));
    errorsSeen = np.zeros((7, 2, 4));
    barsImagined = np.zeros((7,4));
    errorsImagined = np.zeros((7, 2, 4));

    normbarsSeen = np.zeros((7,4));
    normerrorsSeen = np.zeros((7, 2, 4));
    normbarsImagined = np.zeros((7,4));
    normerrorsImagined = np.zeros((7, 2, 4));

    sheetIdxs =  [0, 1, 3, 4, 2, 5];
    numclasses = [200, 50, 200, 50, 50, 50];
    for i in range(len(sheetIdxs)):
        sheet = wb.sheet_by_index(sheetIdxs[i]);
        for j in range(4):
            barsSeen[i, j] = sheet.cell_value(1+j, 6);
            normbarsSeen[i, j] = normalizeAccuracies(sheet.cell_value(1+j, 6), numclasses[i]);

            minValue = 0;
            maxValue = 0;
            for k in range(4):
                value = sheet.cell_value(1+j, k+1) - barsSeen[i, j];
                if k == 0:
                    minValue = value;
                    maxValue = value;
                else:
                    minValue = min(minValue, value);
                    maxValue = max(maxValue, value);
            errorsSeen[i, 0, j] = abs(minValue);
            errorsSeen[i, 1, j] = abs(maxValue);

            minValue = 0;
            maxValue = 0;
            for k in range(4):
                value = normalizeAccuracies(sheet.cell_value(1+j, k+1), numclasses[i]) - normbarsSeen[i, j];
                if k == 0:
                    minValue = value;
                    maxValue = value;
                else:
                    minValue = min(minValue, value);
                    maxValue = max(maxValue, value);
            normerrorsSeen[i, 0, j] = abs(minValue);
            normerrorsSeen[i, 1, j] = abs(maxValue);


        for j in range(4):
            barsImagined[i, j] = sheet.cell_value(7+j, 6);
            normbarsImagined[i, j] = normalizeAccuracies(sheet.cell_value(7+j, 6), numclasses[i]);

            minValue = 0;
            maxValue = 0;
            for k in range(4):
                value = sheet.cell_value(7+j, k+1) - barsImagined[i, j];
                if k == 0:
                    minValue = value;
                    maxValue = value;
                else:
                    minValue = min(minValue, value);
                    maxValue = max(maxValue, value);
            errorsImagined[i, 0, j] = abs(minValue);
            errorsImagined[i, 1, j] = abs(maxValue);

            minValue = 0;
            maxValue = 0;
            for k in range(4):
                value = normalizeAccuracies(sheet.cell_value(7+j, k+1), numclasses[i]) - normbarsImagined[i, j];
                if k == 0:
                    minValue = value;
                    maxValue = value;
                else:
                    minValue = min(minValue, value);
                    maxValue = max(maxValue, value);
            normerrorsImagined[i, 0, j] = abs(minValue);
            normerrorsImagined[i, 1, j] = abs(maxValue);
    
    barsSeen[6, 3] = seenAcc;
    barsImagined[6, 3] = imagAcc;

    plotBarGraphCC(barsSeen, errorsSeen, 'Seen image classification correlation ' + layer, os.path.splitext(outputfile)[0] + '_seen.png');
    plt.pause(0.001);
    plotBarGraphCC(barsImagined, errorsImagined, 'Imagined image classification correlation ' + layer, os.path.splitext(outputfile)[0] + '_imagined.png');


def plotResultsCC2(excel5, excel7, outputfile):
    from openpyxl import load_workbook;

    nBars = 3;
    nScores = 4;
    
    barsSeen = np.zeros((nScores, nBars));
    errorsSeen = np.zeros((nScores, 2, nBars));
    barsImagined = np.zeros((nScores, nBars));
    errorsImagined = np.zeros((nScores, 2, nBars));

    sheetIdxs =  [1, 4];
    for p in range(2):
        if p == 0:
           wb = load_workbook(excel7, data_only=True); 
        else:
           wb = load_workbook(excel5, data_only=True); 
        for i in range(len(sheetIdxs)):
            sheet = wb[wb.sheetnames[sheetIdxs[i]]];
            for j in range(nBars):
                barsSeen[i+p*2, j] = sheet.cell(1+j+1, 6+1).value;
    
                if sheet.cell(1+j+1, 7+1).value == 0:
                    errorsSeen[i+p*2, 0, j] = 0;
                    errorsSeen[i+p*2, 1, j] = 0;
                else:
                    errorsSeen[i+p*2, 0, j] = sheet.cell(1+j+1, 8+1).value;
                    errorsSeen[i+p*2, 1, j] = sheet.cell(1+j+1, 8+1).value;
    
            for j in range(nBars):
                barsImagined[i+p*2, j] = sheet.cell(7+j+1, 6+1).value;
    
                if sheet.cell(7+j+1, 7+1).value == 0:
                    errorsImagined[i+p*2, 0, j] = 0;
                    errorsImagined[i+p*2, 1, j] = 0;
                else:
                    errorsImagined[i+p*2, 0, j] = sheet.cell(7+j+1, 8+1).value;
                    errorsImagined[i+p*2, 1, j] = sheet.cell(7+j+1, 8+1).value;
    
    
    plotBarGraphCC2(barsSeen, errorsSeen, 'Seen stimuli classification by correlation rank with 50 classes', os.path.splitext(outputfile)[0] + '_seen.png');
    plt.pause(0.001);
    plotBarGraphCC2(barsImagined, errorsImagined, 'Imagined stimuli classification by correlation rank with 50 classes', os.path.splitext(outputfile)[0] + '_imagined.png');

def plotResultsCC3(excel5, excel7, outputfile):
    from openpyxl import load_workbook;

    nBars = 3;
    nScores = 4;
    
    barsSeen = np.zeros((nScores, nBars));
    errorsSeen = np.zeros((nScores, 2, nBars));
    barsImagined = np.zeros((nScores, nBars));
    errorsImagined = np.zeros((nScores, 2, nBars));

    sheetIdxs =  [4, 5];
    for p in range(2):
        if p == 0:
           wb = load_workbook(excel7, data_only=True); 
        else:
           wb = load_workbook(excel5, data_only=True); 
           
        for i in range(len(sheetIdxs)):
            sheet = wb[wb.sheetnames[sheetIdxs[i]]];
            
            for j in range(nBars):
                barsSeen[i+p*2, j] = sheet.cell(1+j+1, 6+1).value;
    
                if sheet.cell(1+j+1, 7+1) == 0:
                    errorsSeen[i+p*2, 0, j] = 0;
                    errorsSeen[i+p*2, 1, j] = 0;
                else:
                    errorsSeen[i+p*2, 0, j] = sheet.cell(1+j+1, 8+1).value;
                    errorsSeen[i+p*2, 1, j] = sheet.cell(1+j+1, 8+1).value;
    
            for j in range(nBars):
                barsImagined[i+p*2, j] = sheet.cell(7+j+1, 6+1).value;
    
                if sheet.cell(7+j+1, 7+1).value == 0:
                    errorsImagined[i+p*2, 0, j] = 0;
                    errorsImagined[i+p*2, 1, j] = 0;
                else:
                    errorsImagined[i+p*2, 0, j] = sheet.cell(7+j+1, 8+1).value;
                    errorsImagined[i+p*2, 1, j] = sheet.cell(7+j+1, 8+1).value;
    
    plotBarGraphCC3(barsSeen, errorsSeen, 'Seen stimuli classification by correlation rank with 50 classes', os.path.splitext(outputfile)[0] + '_seen.png');
    plt.pause(0.001);
    plotBarGraphCC3(barsImagined, errorsImagined, 'Imagined stimuli classification by correlation rank with 50 classes', os.path.splitext(outputfile)[0] + '_imagined.png');
    

def plotResultsCC4(outputfile):
    def calculateMatrixResult(filename, bars, idx):
        top1, top5, top10 = matrixAccuracy(filename);
        bars[idx, 0] = top1;
        bars[idx, 1] = top5;
        bars[idx, 2] = top10;
        
        return bars;
        
    nBars = 3;
    nScores = 4;
    
    barsSeen = np.zeros((nScores, nBars));
    errorsSeen = np.zeros((nScores, 2, nBars));
    barsImagined = np.zeros((nScores, nBars));
    errorsImagined = np.zeros((nScores, 2, nBars));

    # fc7 
    filename = 'CMs/AverageMatrix_4096_allData_50Classes_mean';
    barsSeen = calculateMatrixResult(filename, barsSeen, 0);
    filename = 'CMs/AverageMatrix_4096_allData_50Classes_ip';
    barsSeen = calculateMatrixResult(filename, barsSeen, 1);

    filename = 'CMs/AverageMatrix_4096_img_allData_50Classes_mean';
    barsImagined = calculateMatrixResult(filename, barsImagined, 0);
    filename = 'CMs/AverageMatrix_4096_img_allData_50Classes_ip';
    barsImagined = calculateMatrixResult(filename, barsImagined, 1);


    # conv5 
    filename = 'CMs/AverageMatrix_25088_allData_50Classes_mean';
    barsSeen = calculateMatrixResult(filename, barsSeen, 2);
    filename = 'CMs/AverageMatrix_25088_allData_50Classes_ip';
    barsSeen = calculateMatrixResult(filename, barsSeen, 3);

    filename = 'CMs/AverageMatrix_25088_img_allData_50Classes_mean';
    barsImagined = calculateMatrixResult(filename, barsImagined, 2);
    filename = 'CMs/AverageMatrix_25088_img_allData_50Classes_ip';
    barsImagined = calculateMatrixResult(filename, barsImagined, 3);

    
    plotBarGraphCC3(barsSeen, errorsSeen, 'Seen stimuli classification based on matrices', os.path.splitext(outputfile)[0] + '_seen.png');
    plt.pause(0.001);
    plotBarGraphCC3(barsImagined, errorsImagined, 'Imagined stimuli classification based on matrices', os.path.splitext(outputfile)[0] + '_imagined.png');

def plotResultsNN(excelfile, outputfile, layer):
    import xlrd;
    
    wb = xlrd.open_workbook(excelfile); 
    barsSeen = np.zeros((4,4));
    errorsSeen = np.zeros((4, 2, 4));
    barsImagined = np.zeros((4,4));
    errorsImagined = np.zeros((4, 2, 4));

    sheetIdxs =  [1, 3, 2, 4];
    for i in range(4):
        sheet = wb.sheet_by_index(sheetIdxs[i]);
        for j in range(4):
            barsSeen[i, j] = sheet.cell_value(1+j, 6);

            minValue = 0;
            maxValue = 0;
            for k in range(4):
                value = sheet.cell_value(1+j, k+1) - barsSeen[i, j];
                if k == 0:
                    minValue = value;
                    maxValue = value;
                else:
                    minValue = min(minValue, value);
                    maxValue = max(maxValue, value);
            errorsSeen[i, 0, j] = abs(minValue);
            errorsSeen[i, 1, j] = abs(maxValue);

        for j in range(4):
            barsImagined[i, j] = sheet.cell_value(7+j, 6);

            minValue = 0;
            maxValue = 0;
            for k in range(4):
                value = sheet.cell_value(7+j, k+1) - barsImagined[i, j];
                if k == 0:
                    minValue = value;
                    maxValue = value;
                else:
                    minValue = min(minValue, value);
                    maxValue = max(maxValue, value);
            errorsImagined[i, 0, j] = abs(minValue);
            errorsImagined[i, 1, j] = abs(maxValue);
    
    plotBarGraphNN(barsSeen, errorsSeen, 'Seen image classification Neural Networks ' + layer, os.path.splitext(outputfile)[0] + '_seen.png');
    plt.pause(0.001);
    plotBarGraphNN(barsImagined, errorsImagined, 'Imagined image classification Neural Networks ' + layer, os.path.splitext(outputfile)[0] + '_imagined.png');

def plotResultsNN2(excel5, excel7, outputfile):
    import xlrd;

    nBars = 3;
    nScores = 4;
    
    barsSeen = np.zeros((nScores, nBars));
    errorsSeen = np.zeros((nScores, 2, nBars));
    barsImagined = np.zeros((nScores, nBars));
    errorsImagined = np.zeros((nScores, 2, nBars));
    

    sheetIdxs =  [3, 4];
    for p in range(2):
        if p == 0:
           wb = xlrd.open_workbook(excel7); 
        else:
           wb = xlrd.open_workbook(excel5); 
        for i in range(len(sheetIdxs)):
            sheet = wb.sheet_by_index(sheetIdxs[i]);
            for j in range(nBars):
                barsSeen[i+p*2, j] = sheet.cell_value(1+j, 6);
    
                if sheet.cell_value(1+j, 7) == 0:
                    errorsSeen[i+p*2, 0, j] = 0;
                    errorsSeen[i+p*2, 1, j] = 0;
                else:
                    errorsSeen[i+p*2, 0, j] = min(sheet.cell_value(1+j, 6), sheet.cell_value(1+j, 8));
                    errorsSeen[i+p*2, 1, j] = sheet.cell_value(1+j, 8);
    
            for j in range(nBars):
                barsImagined[i+p*2, j] = sheet.cell_value(7+j, 6);
    
                if sheet.cell_value(7+j, 7) == 0:
                    errorsImagined[i+p*2, 0, j] = 0;
                    errorsImagined[i+p*2, 1, j] = 0;
                else:
                    errorsImagined[i+p*2, 0, j] = min(sheet.cell_value(7+j, 6), sheet.cell_value(7+j, 8));
                    errorsImagined[i+p*2, 1, j] = sheet.cell_value(7+j, 8);
    
    
    plotBarGraphNN2(barsSeen, errorsSeen, 'Seen stimuli classification by neural networks with 50 classes', os.path.splitext(outputfile)[0] + '_seen.png');
    plt.pause(0.001);
    plotBarGraphNN2(barsImagined, errorsImagined, 'Imagined stimuli classification neural networks with 50 classes', os.path.splitext(outputfile)[0] + '_imagined.png');

def plotResultsNN3(excelNN7, excelCC7, excelNN5, excelCC5, outputfile):
    import xlrd;

    nBars = 3;
    nScores = 4;
    
    barsSeen = np.zeros((nScores, nBars));
    errorsSeen = np.zeros((nScores, 2, nBars));
    barsImagined = np.zeros((nScores, nBars));
    errorsImagined = np.zeros((nScores, 2, nBars));

    for barIdx in range(4):
        if barIdx == 0:
           wb = xlrd.open_workbook(excelNN7);
           sheetN = 4;
        elif barIdx == 1:
           wb = xlrd.open_workbook(excelCC7);
           sheetN = 4;
        elif barIdx == 2:
           wb = xlrd.open_workbook(excelNN5);
           sheetN = 4;
        elif barIdx == 3:
           wb = xlrd.open_workbook(excelCC5);
           sheetN = 4;
           
        sheet = wb.sheet_by_index(sheetN);
        for j in range(nBars):
            barsSeen[barIdx, j] = sheet.cell_value(1+j, 6);
    
            if sheet.cell_value(1+j, 7) == 0:
                errorsSeen[barIdx, 0, j] = 0;
                errorsSeen[barIdx, 1, j] = 0;
            else:
                errorsSeen[barIdx, 0, j] = min(sheet.cell_value(1+j, 6), sheet.cell_value(1+j, 8));
                errorsSeen[barIdx, 1, j] = sheet.cell_value(1+j, 8);
    
        for j in range(nBars):
            barsImagined[barIdx, j] = sheet.cell_value(7+j, 6);
    
            if sheet.cell_value(7+j, 7) == 0:
                errorsImagined[barIdx, 0, j] = 0;
                errorsImagined[barIdx, 1, j] = 0;
            else:
                errorsImagined[barIdx, 0, j] = min(sheet.cell_value(7+j, 6), sheet.cell_value(7+j, 8));
                errorsImagined[barIdx, 1, j] = sheet.cell_value(7+j, 8);

    plotBarGraphNN3(barsSeen, errorsSeen, 'Seen stimuli classification with 50 classes', os.path.splitext(outputfile)[0] + '_seen.png');
    plt.pause(0.001);
    plotBarGraphNN3(barsImagined, errorsImagined, 'Imagined stimuli classification with 50 classes', os.path.splitext(outputfile)[0] + '_imagined.png');

def plotResultsHuge2(excel7, excel5, outputfile, seenAcc7, imagAcc7, seenAcc5, imagAcc5):
    from openpyxl import load_workbook;

    barsSeen = np.zeros((2,2));
    errorsSeen = np.zeros((2, 2, 2));
    barsImagined = np.zeros((2,2));
    errorsImagined = np.zeros((2, 2, 2));

    # FC7 Horikawa accuracies
    barsSeen[1, 0] = seenAcc7;
    barsImagined[1, 0] = imagAcc7;
  
    errorsSeen[1, 0, 0] = 1.57;
    errorsSeen[1, 1, 0] = 1.57;

    errorsSeen[1, 0, 1] = 1.52;
    errorsSeen[1, 1, 1] = 1.52;

    # Conv5 Horikawa accuracies
    barsSeen[1, 1] = seenAcc5;
    barsImagined[1, 1] = imagAcc5;

    errorsImagined[1, 0, 0] = 6.44;
    errorsImagined[1, 1, 0] = 6.44;

    errorsImagined[1, 0, 1] = 2.35;
    errorsImagined[1, 1, 1] = 2.35;

    sheetIdxs =  [0];
    for i in range(len(sheetIdxs)):
        for j in range(2):
            if j == 0:
               wb = load_workbook(excel7, data_only=True); 
            else:
               wb = load_workbook(excel5, data_only=True); 
               
            sheet = wb.active;

            barsSeen[i, j] = sheet.cell(5, 7).value;
            if sheet.cell(5, 9).value == 0:     
                errorsSeen[i, 0, j] = 0;
                errorsSeen[i, 1, j] = 0;
            else:
                errorsSeen[i, 0, j] = sheet.cell(5, 9).value;
                errorsSeen[i, 1, j] = sheet.cell(5, 9).value;

            barsImagined[i, j] = sheet.cell(11, 7).value;

            if sheet.cell(10, 8).value == 0:    
                errorsImagined[i, 0, j] = 0;
                errorsImagined[i, 1, j] = 0;
            else:
                errorsImagined[i, 0, j] = sheet.cell(11, 9).value;
                errorsImagined[i, 1, j] = sheet.cell(11, 9).value;

    
    plotBarGraphHuge2(barsSeen, errorsSeen, 'Seen image classification correlation 14018 classes ', os.path.splitext(outputfile)[0] + '_seen.png');
    plt.pause(0.001);
    plotBarGraphHuge2(barsImagined, errorsImagined, 'Imagined image classification correlation 14018 classes ', os.path.splitext(outputfile)[0] + '_imagined.png');

def plotResultsHuge3(excel7, excel5, outputfile, seenAcc7, imagAcc7, seenAcc5, imagAcc5):
    import xlrd;
    
    barsSeen = np.zeros((3,2));
    errorsSeen = np.zeros((3, 2, 2));
    barsImagined = np.zeros((3,2));
    errorsImagined = np.zeros((3, 2, 2));

    # FC7 Horikawa accuracies
    barsSeen[2, 0] = seenAcc7;
    barsImagined[2, 0] = imagAcc7;

    # Conv5 Horikawa accuracies
    barsSeen[2, 1] = seenAcc5;
    barsImagined[2, 1] = imagAcc5;

    sheetIdxs =  [0, 1];
    for i in range(len(sheetIdxs)):
        for j in range(2):
            if j == 0:
               wb = xlrd.open_workbook(excel7); 
            else:
               wb = xlrd.open_workbook(excel5); 
               
            sheet = wb.sheet_by_index(sheetIdxs[i]);
    
            barsSeen[i, j] = sheet.cell_value(4, 6);
    
            if sheet.cell_value(4, 7) == 0:
                errorsSeen[i, 0, j] = 0;
                errorsSeen[i, 1, j] = 0;
            else:
                errorsSeen[i, 0, j] = sheet.cell_value(4, 8);
                errorsSeen[i, 1, j] = sheet.cell_value(4, 8);
    
            barsImagined[i, j] = sheet.cell_value(10, 6);
    
            if sheet.cell_value(10, 7) == 0:
               errorsImagined[i, 0, j] = 0;
               errorsImagined[i, 1, j] = 0;
            else:
               errorsImagined[i, 0, j] = sheet.cell_value(10, 8);
               errorsImagined[i, 1, j] = sheet.cell_value(10, 8);

    
    plotBarGraphHuge2(barsSeen, errorsSeen, 'Seen image classification correlation 14018 classes ', os.path.splitext(outputfile)[0] + '_seen.png');
    plt.pause(0.001);
    plotBarGraphHuge2(barsImagined, errorsImagined, 'Imagined image classification correlation 14018 classes ', os.path.splitext(outputfile)[0] + '_imagined.png');

def plotResultsHuge(excelfile, outputfile, layer, seenAcc, imagAcc):
    import xlrd;
    
    wb = xlrd.open_workbook(excelfile); 
    barsSeen = np.zeros((3,4));
    errorsSeen = np.zeros((3, 2, 4));
    barsImagined = np.zeros((3,4));
    errorsImagined = np.zeros((3, 2, 4));

    sheetIdxs =  [0, 1];
    for i in range(len(sheetIdxs)):
        sheet = wb.sheet_by_index(sheetIdxs[i]);
        for j in range(4):
            barsSeen[i, j] = sheet.cell_value(1+j, 6);

            minValue = 0;
            maxValue = 0;
            for k in range(4):
                value = sheet.cell_value(1+j, k+1) - barsSeen[i, j];
                if k == 0:
                    minValue = value;
                    maxValue = value;
                else:
                    minValue = min(minValue, value);
                    maxValue = max(maxValue, value);
            errorsSeen[i, 0, j] = abs(minValue);
            errorsSeen[i, 1, j] = abs(maxValue);

        for j in range(4):
            barsImagined[i, j] = sheet.cell_value(7+j, 6);

            minValue = 0;
            maxValue = 0;
            for k in range(4):
                value = sheet.cell_value(7+j, k+1) - barsImagined[i, j];
                if k == 0:
                    minValue = value;
                    maxValue = value;
                else:
                    minValue = min(minValue, value);
                    maxValue = max(maxValue, value);
            errorsImagined[i, 0, j] = abs(minValue);
            errorsImagined[i, 1, j] = abs(maxValue);
    
    barsSeen[2, 3] = seenAcc;
    barsImagined[2, 3] = imagAcc;

    plotBarGraphHuge(barsSeen, errorsSeen, 'Seen image classification correlation 15225 classes' + layer, os.path.splitext(outputfile)[0] + '_seen.png');
    plt.pause(0.001);
    plotBarGraphHuge(barsImagined, errorsImagined, 'Imagined image classification correlation 15225 classes' + layer, os.path.splitext(outputfile)[0] + '_imagined.png');


def plotResultsMulti(excelfile, outputfile):
    import xlrd;
    
    wb = xlrd.open_workbook(excelfile); 
    barsSeen = np.zeros((2,4));
    errorsSeen = np.zeros((2, 2, 4));

    sheet = wb.sheet_by_index(0);
    i = 0;
    for j in range(4):
        barsSeen[i, j] = sheet.cell_value(1+j, 6);

        minValue = 0;
        maxValue = 0;
        for k in range(4):
            value = sheet.cell_value(1+j, k+1) - barsSeen[i, j];
            if k == 0:
                minValue = value;
                maxValue = value;
            else:
                minValue = min(minValue, value);
                maxValue = max(maxValue, value);
        errorsSeen[i, 0, j] = abs(minValue);
        errorsSeen[i, 1, j] = abs(maxValue);

    i = 1;
    for j in range(4):
        barsSeen[i, j] = sheet.cell_value(7+j, 6);

        minValue = 0;
        maxValue = 0;
        for k in range(4):
            value = sheet.cell_value(7+j, k+1) - barsSeen[i, j];
            if k == 0:
                minValue = value;
                maxValue = value;
            else:
                minValue = min(minValue, value);
                maxValue = max(maxValue, value);
        errorsSeen[i, 0, j] = abs(minValue);
        errorsSeen[i, 1, j] = abs(maxValue);
    
    plotBarGraphMulti(barsSeen, errorsSeen, 'Seen image classification LOO 50 classes', os.path.splitext(outputfile)[0] + '_seen.png');

#plotResultsCC4('correlation_25088_CM.png');