import os;
from data_handler import data_handler;
from openpyxl.workbook import Workbook;
from datetime import datetime;

import numpy as np;
from DNNUtils import makeCoders, createFeatures;
from resultUtils import getColumn, printResults;
import torch;
from varUtils import fmri_dir, img_dir, confusionMatrixDir, resultDir;
from cnn2 import run_decode_cnn;
import xlwings;




def openExcelFile(excelFile):
    excel_app = xlwings.App(visible=False);

    excel_book = excel_app.books.open(excelFile);
    excel_book.save();
    excel_book.close();

    excel_app.quit()


def prototypeTrain(tipo=1):
    if True:
       subjects = [
            'Subject4'
       ];

       subjects = [
            'Subject1',
            'Subject2',
            'Subject3',
            'Subject4',
            'Subject5'
       ];

    roi = 'ROI_VC';
    if tipo == 1:
       classification, encoder, decoder = makeCoders();
    else:
       classification, encoder, decoder = makeCoders(0);
    
    wb = Workbook();
    ws = wb['Sheet'];
    ws.title = "200 classes all data";
    ws = wb.create_sheet("50 classes all data");
    ws = wb.create_sheet("50 classes all data image p.");

    ws = wb.create_sheet("200 classes mean data");
    ws = wb.create_sheet("50 classes mean data");
    ws = wb.create_sheet("50 classes mean data image p.");
    
    if tipo == 1:
        sizeSuffix = '4096_';
    else:
        sizeSuffix = '25088_';
    
    
    for ws in wb.worksheets:
        ws['B1'] = 'Subject 1';
        ws['C1'] = 'Subject 2';
        ws['D1'] = 'Subject 3';
        ws['E1'] = 'Subject 4';
        ws['F1'] = 'Subject 5';
        ws['G1'] = 'Mean';
        ws['H1'] = 'Standard Deviation';
        ws['I1'] = 'CI';
       
        # average
        ws['G2'] = '=AVERAGE(B2:F2)';
        ws['G3'] = '=AVERAGE(B3:F3)';
        ws['G4'] = '=AVERAGE(B4:F4)';
        ws['G5'] = '=AVERAGE(B5:F5)';

        ws['G8'] = '=AVERAGE(B8:F8)';
        ws['G9'] = '=AVERAGE(B9:F9)';
        ws['G10'] = '=AVERAGE(B10:F10)';
        ws['G11'] = '=AVERAGE(B11:F11)';

        # standard deviation
        ws['H2'] = '=STDEV(B2:F2)';
        ws['H3'] = '=STDEV(B3:F3)';
        ws['H4'] = '=STDEV(B4:F4)';
        ws['H5'] = '=STDEV(B5:F5)';

        ws['H8'] = '=STDEV(B8:F8)';
        ws['H9'] = '=STDEV(B9:F9)';
        ws['H10'] = '=STDEV(B10:F10)';
        ws['H11'] = '=STDEV(B11:F11)';

        # confidence interval
        ws['I2'] = '=CONFIDENCE(0.05, H2, 5)';
        ws['I3'] = '=CONFIDENCE(0.05, H3, 5)';
        ws['I4'] = '=CONFIDENCE(0.05, H4, 5)';
        ws['I5'] = '=CONFIDENCE(0.05, H5, 5)';

        ws['I8'] = '=CONFIDENCE(0.05, H8, 5)';
        ws['I9'] = '=CONFIDENCE(0.05, H9, 5)';
        ws['I10'] = '=CONFIDENCE(0.05, H10, 5)';
        ws['I11'] = '=CONFIDENCE(0.05, H11, 5)';

        # line names
        ws['A2'] = 'Top 1  Accuracy';
        ws['A3'] = 'Top 5  Accuracy';
        ws['A4'] = 'Top 10 Accuracy';
        ws['A5'] = 'Kamitani  Score';
        
        ws['A7'] = 'Imagined data';
        
        ws['A8'] = 'Top 1  Accuracy';
        ws['A9'] = 'Top 5  Accuracy';
        ws['A10']= 'Top 10 Accuracy';
        ws['A11']= 'Kamitani  Score';


    for sbj in subjects:
        # load prep file generated from bdPy files from Kamitani Lab
        print('');
        print('Processando %s' % sbj);
        print('Loading data from disk.');
        subjectFile = '%s/%s.h5' % (fmri_dir, sbj);
        handler = data_handler(subjectFile);
        x_train, x_test, x_test_avg, x_imag, x_imag_avg = handler.get_data(roi = roi, imag_data = 1, unityNormalization = 1);
        #return x_train, x_test, x_test_avg, x_imag, x_imag_avg;

        lbl_train, lbl_test, lbl_imag = handler.get_files(imag_data = 1);
        _, lbl_test_idx, lbl_imag_idx = handler.get_indices(imag_data = 1);
        
        
        lbl_test_avg = [];
        lbl_imag_avg = [];
        for i in range(50):
            idx = np.where(lbl_test_idx == i)[0][0];
            lbl_test_avg.append(lbl_test[idx]);
            
            idx = np.where(lbl_imag_idx == i)[0][0];
            lbl_imag_avg.append(lbl_imag[idx]);

        print('Min x_train : ', np.min(x_train));
        print('Max x_train : ', np.max(x_train));

        print('Min x_test : ', np.min(x_test));
        print('Max x_test : ', np.max(x_test));

        print('Min x_imag : ', np.min(x_imag));
        print('Max x_imag : ', np.max(x_imag));
        
        
        decode_id = 'teste';
        testData = x_test;
        testLabels = lbl_test;
        
        featuresTrain, featuresTest = createFeatures(encoder, tipo, img_dir, lbl_train, testLabels);
        y_train = torch.cat(featuresTrain);
        y_test = torch.cat(featuresTest);
        
        column = getColumn(sbj);

        minTreino = np.min(y_train.cpu().data.numpy());
        maxTreino = np.max(y_train.cpu().data.numpy());

        pred_y, train_pred, transNet, lossCurve = run_decode_cnn(decode_id, x_train, y_train, testData, y_test, minTreino, maxTreino);

        if 1:
            print('Accuracy with all test data');
            ws = wb.worksheets[0];
            printResults(ws, column, 1, 'Accuracy 200 classes (prototype mean)', pred_y, testLabels, 1, tipo);
            
            ws = wb.worksheets[1];
            printResults(ws, column, 1, 'Accuracy 50 classes (prototype mean)', pred_y, testLabels, 2, tipo, CMFile = confusionMatrixDir + '%s_%sallData_50Classes_Mean' % (sbj, sizeSuffix));
    
            ws = wb.worksheets[2];
            printResults(ws, column, 1, 'Accuracy 50 classes (target image)', pred_y, testLabels, 3, tipo, CMFile = confusionMatrixDir + '%s_%sallData_50Classes_IP' % (sbj, sizeSuffix));

        print('');
        print('Accuracy with average test data');
        testData = x_test_avg;
        testLabels = lbl_test_avg;
        
        testData = torch.from_numpy(testData).float(); 
        pred_y = [];
        for i in range(testData.shape[0]):
            prediction = transNet(testData[i].cuda());
            pred_y.append(prediction);
        
        ws = wb.worksheets[3];
        printResults(ws, column, 1, 'Accuracy 200 classes (prototype mean)', pred_y, testLabels, 1, tipo);

        ws = wb.worksheets[4];
        printResults(ws, column, 1, 'Accuracy 50 classes (prototype mean)', pred_y, testLabels, 2, tipo, CMFile = confusionMatrixDir + '%s_%smeanData_50Classes_Mean' % (sbj, sizeSuffix));

        ws = wb.worksheets[5];
        printResults(ws, column, 1, 'Accuracy 50 classes (target image)', pred_y, testLabels, 3, tipo, CMFile = confusionMatrixDir + '%s_%smeanData_50Classes_IP' % (sbj, sizeSuffix));

        
        if 1:
            testData = x_imag;
            testLabels = lbl_imag;
            
            testData = torch.from_numpy(testData).float(); 
            pred_y = [];
            for i in range(testData.shape[0]):
                prediction = transNet(testData[i].cuda());
                pred_y.append(prediction);
                
            print('Accuracy with all imag data');
    
            ws = wb.worksheets[0];
            printResults(ws, column, 7, 'Accuracy 200 classes (prototype mean)', pred_y, testLabels, 1, tipo);
            
            ws = wb.worksheets[1];
            printResults(ws, column, 7, 'Accuracy 50 classes (prototype mean)', pred_y, testLabels, 2, tipo, CMFile = confusionMatrixDir + '%s_%simg_allData_50Classes_Mean' % (sbj, sizeSuffix));
    
            ws = wb.worksheets[2];
            printResults(ws, column, 7, 'Accuracy 50 classes (target image)', pred_y, testLabels, 3, tipo, CMFile = confusionMatrixDir + '%s_%simg_allData_50Classes_IP' % (sbj, sizeSuffix));

        print('');
        print('Accuracy with average imag data');
        testData = x_imag_avg;
        testLabels = lbl_imag_avg;
        
        testData = torch.from_numpy(testData).float(); 
        pred_y = [];
        for i in range(testData.shape[0]):
            prediction = transNet(testData[i].cuda());
            pred_y.append(prediction);
        
        
        ws = wb.worksheets[3];
        printResults(ws, column, 7, 'Accuracy 200 classes (prototype mean)', pred_y, testLabels, 1, tipo);

        ws = wb.worksheets[4];
        printResults(ws, column, 7, 'Accuracy 50 classes (prototype mean)', pred_y, testLabels, 2, tipo, CMFile = confusionMatrixDir + '%s_%simg_meanData_50Classes_Mean' % (sbj, sizeSuffix));

        ws = wb.worksheets[5];
        printResults(ws, column, 7, 'Accuracy 50 classes (target image)', pred_y, testLabels, 3, tipo, CMFile = confusionMatrixDir + '%s_%simg_meanData_50Classes_IP' % (sbj, sizeSuffix));

    timestamp = datetime.now().strftime('%Y%m%d%H%M%S');
    if tipo == 1:
        timestamp = '4096_' + timestamp;
    else:
        timestamp = '25088_' + timestamp;
    filename = os.path.join(resultDir, 'resultados_' + timestamp + '.xlsx');
    wb.save(filename = filename);

    # opening and saving excel files as openpyxl does not calculate formulae. 
    # Need excel installed
    openExcelFile(filename);
    return filename;

 
def prototypeTrainHuge(tipo=1, mean_volume = True, single_volume = 0):
    
    if True:
       subjects = [
            'Subject3'
       ];

       subjects = [
            'Subject1',
            'Subject2',
            'Subject3',
            'Subject4',
            'Subject5'
       ];

    roi = 'ROI_VC';
    if tipo == 1:
       classification, encoder, decoder = makeCoders();
    else:
       classification, encoder, decoder = makeCoders(0);
    
    wb = Workbook();
    ws = wb['Sheet'];
    ws.title = "Huge classes mean data";
    
    for ws in wb.worksheets:
        ws['B1'] = 'Subject 1';
        ws['C1'] = 'Subject 2';
        ws['D1'] = 'Subject 3';
        ws['E1'] = 'Subject 4';
        ws['F1'] = 'Subject 5';
        ws['G1'] = 'Mean';
        ws['H1'] = 'Standard Deviation';
        ws['I1'] = 'CI';
       
        # average
        ws['G2'] = '=AVERAGE(B2:F2)';
        ws['G3'] = '=AVERAGE(B3:F3)';
        ws['G4'] = '=AVERAGE(B4:F4)';
        ws['G5'] = '=AVERAGE(B5:F5)';

        ws['G8'] = '=AVERAGE(B8:F8)';
        ws['G9'] = '=AVERAGE(B9:F9)';
        ws['G10'] = '=AVERAGE(B10:F10)';
        ws['G11'] = '=AVERAGE(B11:F11)';

        # standard deviation
        ws['H2'] = '=STDEV(B2:F2)';
        ws['H3'] = '=STDEV(B3:F3)';
        ws['H4'] = '=STDEV(B4:F4)';
        ws['H5'] = '=STDEV(B5:F5)';

        ws['H8'] = '=STDEV(B8:F8)';
        ws['H9'] = '=STDEV(B9:F9)';
        ws['H10'] = '=STDEV(B10:F10)';
        ws['H11'] = '=STDEV(B11:F11)';

        # confidence interval
        ws['I2'] = '=CONFIDENCE(0.05, H2, 5)';
        ws['I3'] = '=CONFIDENCE(0.05, H3, 5)';
        ws['I4'] = '=CONFIDENCE(0.05, H4, 5)';
        ws['I5'] = '=CONFIDENCE(0.05, H5, 5)';

        ws['I8'] = '=CONFIDENCE(0.05, H8, 5)';
        ws['I9'] = '=CONFIDENCE(0.05, H9, 5)';
        ws['I10'] = '=CONFIDENCE(0.05, H10, 5)';
        ws['I11'] = '=CONFIDENCE(0.05, H11, 5)';

        # line names
        ws['A2'] = 'Top 1  Accuracy';
        ws['A3'] = 'Top 5  Accuracy';
        ws['A4'] = 'Top 10 Accuracy';
        ws['A5'] = 'Kamitani  Score';
        
        ws['A7'] = 'Imagined data';
        
        ws['A8'] = 'Top 1  Accuracy';
        ws['A9'] = 'Top 5  Accuracy';
        ws['A10']= 'Top 10 Accuracy';
        ws['A11']= 'Kamitani  Score';

    for sbj in subjects:
        # load prep file generated from bdPy files from Kamitani Lab
        print('');
        print('Processando %s' % sbj);
        print('Loading data from disk.');
        subjectFile = '%s/%s.h5' % (fmri_dir, sbj);
        handler = data_handler(subjectFile);
        x_train, x_test, x_test_avg, x_imag, x_imag_avg = handler.get_data(roi = roi, unityNormalization = 1, imag_data = 1);
        lbl_train, lbl_test, lbl_imag = handler.get_files(imag_data = 1);
        _, lbl_test_idx, lbl_imag_idx = handler.get_indices(imag_data = 1);
        
        lbl_test_avg = [];
        lbl_imag_avg = [];
        for i in range(50):
            idx = np.where(lbl_test_idx == i)[0][0];
            lbl_test_avg.append(lbl_test[idx]);
            
            idx = np.where(lbl_imag_idx == i)[0][0];
            lbl_imag_avg.append(lbl_imag[idx]);


        print('Min x_train : ', np.min(x_train));
        print('Max x_train : ', np.max(x_train));

        print('Min x_test : ', np.min(x_test));
        print('Max x_test : ', np.max(x_test));

        print('Min x_imag : ', np.min(x_imag));
        print('Max x_imag : ', np.max(x_imag));
        
        decode_id = 'teste';
        testData = x_test;
        testLabels = lbl_test;
        
        featuresTrain, featuresTest = createFeatures(encoder, tipo, img_dir, lbl_train, testLabels);
        y_train = torch.cat(featuresTrain);
        y_test = torch.cat(featuresTest);
        
        column = getColumn(sbj);

        minTreino = np.min(y_train.cpu().data.numpy());
        maxTreino = np.max(y_train.cpu().data.numpy());

        pred_y, train_pred, transNet, lossCurve = run_decode_cnn(decode_id, x_train, y_train, testData, y_test, minTreino, maxTreino);


        print('Accuracy with all test data');
        if single_volume:
           ws = wb.worksheets[0];
           printResults(ws, column, 1, 'Accuracy Huge classes (prototype mean)', pred_y, testLabels, 5, tipo);
        
        print('');
        print('Accuracy with average test data');
        testData = x_test_avg;
        testLabels = lbl_test_avg;
        
        testData = torch.from_numpy(testData).float(); 
        pred_y = [];
        for i in range(testData.shape[0]):
            prediction = transNet(testData[i].cuda());
            pred_y.append(prediction);
        
        if mean_volume:
           ws = wb.worksheets[0];
           printResults(ws, column, 1, 'Accuracy Huge classes (prototype mean)', pred_y, testLabels, 5, tipo);

        testData = x_imag;
        testLabels = lbl_imag;
        
        testData = torch.from_numpy(testData).float(); 
        pred_y = [];
        for i in range(testData.shape[0]):
            prediction = transNet(testData[i].cuda());
            pred_y.append(prediction);
            
        print('Accuracy with all imag data');

        if single_volume:
           ws = wb.worksheets[0];
           printResults(ws, column, 7, 'Accuracy Huge classes (prototype mean)', pred_y, testLabels, 5, tipo);
        
        print('');
        print('Accuracy with average imag data');
        testData = x_imag_avg;
        testLabels = lbl_imag_avg;
        
        testData = torch.from_numpy(testData).float(); 
        pred_y = [];
        for i in range(testData.shape[0]):
            prediction = transNet(testData[i].cuda());
            pred_y.append(prediction);
        
        
        if mean_volume:
           ws = wb.worksheets[0];
           printResults(ws, column, 7, 'Accuracy Huge classes (prototype mean)', pred_y, testLabels, 5, tipo);

    timestamp = datetime.now().strftime('%Y%m%d%H%M%S');
    if tipo == 1:
        timestamp = '4096_' + timestamp;
    else:
        timestamp = '25088_' + timestamp;

    filename = os.path.join(resultDir, 'resultados_huge_' + timestamp + '.xlsx');
    wb.save(filename = filename);

    # opening and saving excel files as openpyxl does not calculate formulae. 
    # Need excel installed
    openExcelFile(filename);
    
    return filename;
