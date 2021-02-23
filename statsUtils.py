from listUtils import loadList;
import numpy as np;
from scipy import stats;
from varUtils import  img_dir;

def one_sample_one_tailed2(sample_data, popmean, alpha=0.05, alternative='greater'):
    t, p = stats.ttest_1samp(sample_data, popmean)
    showMessage = False;
    if showMessage:
        print ('t:',t)
        print ('p:',p)
        if alternative == 'greater' and (p/2 < alpha) and t > 0:
            print ('Reject Null Hypothesis for greater-than test')
        if alternative == 'less' and (p/2 < alpha) and t < 0:
            print ('Reject Null Hypothesis for less-thane test')
    return p/2;

def one_sample_one_tailed(sample_data, popmean, alpha=0.05, alternative='greater'):
    t, p = stats.wilcoxon(np.array(sample_data) - popmean, alternative = alternative);
    showMessage = False;
    if showMessage:
        print ('t:',t)
        print ('p:',p)
        if alternative == 'greater' and (p/2 < alpha) and t > 0:
            print ('Reject Null Hypothesis for greater-than test')
        if alternative == 'less' and (p/2 < alpha) and t < 0:
            print ('Reject Null Hypothesis for less-thane test')
    return p;

def verifyDifference(excel5, excel7, horikawa):
    import xlrd;

    wb5 = xlrd.open_workbook(excel5); 
    wb7 = xlrd.open_workbook(excel7); 
    wbh = xlrd.open_workbook(horikawa); 
    
    sheet7 = wb7.sheet_by_index(1);
    sheet5 = wb5.sheet_by_index(1);

    sheeth = wbh.sheet_by_index(3);
    # Teste Visto fc7
    linha = 4;
    sample_data = [];
    for i in range(5):
        sample_data.append(sheet7.cell_value(linha, i+1));

    sample_datah = [];
    for i in range(5):
        sample_datah.append(sheeth.cell_value(linha, i+1));
    
    z, p = stats.wilcoxon(sample_data, sample_datah, alternative='greater');
    print('FC7 Visto');
    print('z = ', z);
    print('p = ', p);

    # Teste Imaginado fc7
    linha = 10;
    sample_data = [];
    for i in range(5):
        sample_data.append(sheet7.cell_value(linha, i+1));

    sample_datah = [];
    for i in range(5):
        sample_datah.append(sheeth.cell_value(linha, i+1));
    
    z, p = stats.wilcoxon(sample_data, sample_datah, alternative='greater');
    print('FC7 Imaginado');
    print('z = ', z);
    print('p = ', p);
    
    sheeth = wbh.sheet_by_index(2);
    # Teste Visto conv5
    linha = 4;
    sample_data = [];
    for i in range(5):
        sample_data.append(sheet5.cell_value(linha, i+1));

    sample_datah = [];
    for i in range(5):
        sample_datah.append(sheeth.cell_value(linha, i+1));
    
    z, p = stats.wilcoxon(sample_data, sample_datah, alternative='greater');
    print('Conv5 Visto' );
    print('Z = ', z);
    print('P = ', p);

    # Teste Imaginado conv5
    linha = 10;
    sample_data = [];
    for i in range(5):
        sample_data.append(sheet5.cell_value(linha, i+1));

    sample_datah = [];
    for i in range(5):
        sample_datah.append(sheeth.cell_value(linha, i+1));
    
    z, p = stats.wilcoxon(sample_data, sample_datah, alternative='greater');
    print('Conv5 Imaginado' );
    print('Z = ', z);
    print('P = ', p);

def verifyDifference2(excel5, excel7):
    import xlrd;

    wb5 = xlrd.open_workbook(excel5); 
    wb7 = xlrd.open_workbook(excel7); 
    
    sheetAP = wb7.sheet_by_index(4);
    sheetIP = wb7.sheet_by_index(5);

    # Teste fc7 increase
    for i in range(3):
        sample_dataAP = [];
        sample_dataIP = [];
        for s in range(5):
            sample_dataAP.append(sheetAP.cell_value(1+i, s+1));
            sample_dataIP.append(sheetIP.cell_value(1+i, s+1));
    
        z, p = stats.wilcoxon(sample_dataIP, sample_dataAP, alternative='greater');
        #z, p = stats.ttest_rel(sample_dataIP, sample_dataAP);
        #print(sample_dataAP);
        #print(sample_dataIP);
        print('FC7 Increase');
        if i == 0:
            print('Top 1')
        elif i == 1:
            print('Top 5')
        elif i == 2:
            print('Top 10')
        print('z = ', z);
        print('p = ', p);
        print('');

    sheetAP = wb5.sheet_by_index(4);
    sheetIP = wb5.sheet_by_index(5);

    # Teste conv5 increase
    for i in range(3):
        sample_dataAP = [];
        sample_dataIP = [];
        for s in range(5):
            sample_dataAP.append(sheetAP.cell_value(1+i, s+1));
            sample_dataIP.append(sheetIP.cell_value(1+i, s+1));
    
        z, p = stats.wilcoxon(sample_dataIP, sample_dataAP, alternative='greater');
        #z, p = stats.ttest_rel(sample_dataIP, sample_dataAP);
        #print(sample_dataAP);
        #print(sample_dataIP);
        print('Conv5 Increase');
        if i == 0:
            print('Top 1')
        elif i == 1:
            print('Top 5')
        elif i == 2:
            print('Top 10')
        print('z = ', z);
        print('p = ', p);
        print('');
    
def verifySignificance(excelfile, sheetNumber, inicio=1, Kamitani = False):
    from openpyxl import load_workbook;
    
    showMessage = False;     
    wb = load_workbook(excelfile, data_only=True); 
    sheet = wb[wb.sheetnames[sheetNumber]];

    if Kamitani:
        if showMessage:
            print('');
            print('Kamitani Score');
        popmean = 50;
        linha = 3;
        sample_data = [];
        for i in range(5):
            sample_data.append(sheet.cell(inicio+linha+1, i+1+1).value);
        
        if showMessage:
            print('Mean :', sheet.cell(inicio+linha+1, 6+1).value);
            print('SD   :', sheet.cell(inicio+linha+1, 7+1).value);
            print('CI   :', sheet.cell(inicio+linha+1, 8+1).value);
        p = one_sample_one_tailed(sample_data, popmean);
        return p;
        
    else:
        if showMessage:
            print('');
            print('Top 1 accuracy');
        popmean = 1/50;
        linha = 0;
        sample_data = [];
        for i in range(5):
            sample_data.append(sheet.cell(inicio+linha+1, i+1+1).value);

        if showMessage:
            print('Mean :', sheet.cell(inicio+linha+1, 6+1).value);
            print('SD   :', sheet.cell(inicio+linha+1, 7+1).value);
            print('CI   :', sheet.cell(inicio+linha+1, 8+1).value);
        p1 = one_sample_one_tailed(sample_data, popmean);
        
        if showMessage:
            print('');
            print('Top 5 accuracy');
        popmean = 5/50;
        linha = 1;
        sample_data = [];
        for i in range(5):
            sample_data.append(sheet.cell(inicio+linha+1, i+1+1).value);

        if showMessage:
            print('Mean :', sheet.cell(inicio+linha+1, 6+1).vakue);
            print('SD   :', sheet.cell(inicio+linha+1, 7+1).value);
            print('CI   :', sheet.cell(inicio+linha+1, 8+1).value);
        p5 = one_sample_one_tailed(sample_data, popmean);
    
        if showMessage:
            print('');
            print('Top 10 accuracy');
        popmean = 10/50;
        linha = 2;
        sample_data = [];
        for i in range(5):
            sample_data.append(sheet.cell(inicio+linha+1, i+1+1).value);

        if showMessage:
            print('Mean :', sheet.cell(inicio+linha+1, 6+1).value);
            print('SD   :', sheet.cell(inicio+linha+1, 7+1).value);
            print('CI   :', sheet.cell(inicio+linha+1, 8+1).value);
        
        p10 = one_sample_one_tailed(sample_data, popmean);
        
        print('-----------------------------------------------------');
        
        return p1, p5, p10;

def calculateProximity(class1, class2):
    from nltk.corpus import wordnet as wn;
    a = wn.synset_from_pos_and_offset('n', int(class1[1:]));
    b = wn.synset_from_pos_and_offset('n', int(class2[1:]));
    return a.lch_similarity(b);

def verifiyDistance():
    trainClasses = loadList('%s/train.txt' % img_dir);
    proximity = {};
    for i in range(len(trainClasses)):
        baseClass = trainClasses[i];
        for j in range(i+1, len(trainClasses)):
            if calculateProximity(baseClass, trainClasses[j]) > 2:
               if baseClass not in proximity:
                   proximity[baseClass] = [];
               proximity[baseClass].append(trainClasses[j]);
    print(proximity);        
